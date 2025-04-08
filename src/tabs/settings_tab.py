import os
import json
import logging
import pandas as pd

from PyQt5.QtWidgets import (
    QWidget, QMessageBox, QFileDialog, QComboBox, QSpinBox,
    QPushButton, QCheckBox, QVBoxLayout, QLabel
)
from PyQt5 import uic


class SettingsTab(QWidget):
    def __init__(self, main_window, firebase_manager):
        super().__init__(main_window)
        # Сохраняем ссылку на главное окно под другим именем
        self.main_window = main_window  
        self.firebase_manager = firebase_manager

        # Создаем интерфейс настроек программно (без отдельного .ui файла)
        layout = QVBoxLayout(self)

        # 1. Интервал автообновления данных
        self.refreshIntervalSpinBox = QSpinBox(self)
        self.refreshIntervalSpinBox.setMinimum(1)
        self.refreshIntervalSpinBox.setMaximum(3600)
        self.refreshIntervalSpinBox.setValue(60)
        layout.addWidget(QLabel("Интервал автообновления (сек):"))
        layout.addWidget(self.refreshIntervalSpinBox)

        # 2. Выбор таблицы для экспорта данных
        self.exportTableComboBox = QComboBox(self)
        self.exportTableComboBox.addItems(["Cases", "Proformas", "Clients", "Cargo", "Suppliers"])
        layout.addWidget(QLabel("Выбрать таблицу для экспорта:"))
        layout.addWidget(self.exportTableComboBox)

        # 3. Кнопка экспорта в Excel
        self.exportButton = QPushButton("Export to EXCEL", self)
        layout.addWidget(self.exportButton)

        # 4. Дополнительные параметры отображения
        self.autoResizeCheckBox = QCheckBox("Автоматическое изменение размера таблиц", self)
        layout.addWidget(self.autoResizeCheckBox)

        # 5. Кнопка сохранения настроек
        self.saveSettingsButton = QPushButton("Сохранить настройки", self)
        layout.addWidget(self.saveSettingsButton)

        # Проверка наличия необходимых виджетов (на случай, если позже переключимся на ui-файл)
        missing_widgets = []
        if self.refreshIntervalSpinBox is None:
            missing_widgets.append('refreshIntervalSpinBox')
        if self.exportTableComboBox is None:
            missing_widgets.append('exportTableComboBox')
        if self.exportButton is None:
            missing_widgets.append('exportButton')
        if self.autoResizeCheckBox is None:
            missing_widgets.append('autoResizeCheckBox')
        if self.saveSettingsButton is None:
            missing_widgets.append('saveSettingsButton')
        if missing_widgets:
            msg = f"Отсутствуют следующие виджеты на вкладке 'Настройки': {', '.join(missing_widgets)}"
            logging.error(msg)
            QMessageBox.critical(self.main_window, "Ошибка", msg)
            return

        # Подключаем обработчики
        self.exportButton.clicked.connect(self.export_data)
        self.saveSettingsButton.clicked.connect(self.save_settings)

    def export_data(self):
        """
        Экспортирует данные из выбранной таблицы в Excel.
        После экспорта с помощью openpyxl происходит настройка ширины столбцов.
        """
        try:
            table_name = self.exportTableComboBox.currentText().strip().lower()
            table = None

            if table_name == "cargo":
                table_obj = getattr(self.main_window, "cargo_tab", None)
                if table_obj is not None:
                    table = getattr(table_obj, "cargoTable", None)
            elif table_name == "cases":
                table_obj = getattr(self.main_window, "case_numbers_tab", None)
                if table_obj is not None:
                    table = getattr(table_obj, "caseTable", None)
            elif table_name == "proformas":
                table_obj = getattr(self.main_window, "proforma_tab", None)
                if table_obj is not None:
                    table = getattr(table_obj, "proformaTable", None)
            elif table_name == "clients":
                table_obj = getattr(self.main_window, "clients_tab", None)
                if table_obj is not None:
                    table = getattr(table_obj, "clientTable", None)
            elif table_name == "suppliers":
                table_obj = getattr(self.main_window, "suppliers_tab", None)
                if table_obj is not None:
                    table = getattr(table_obj, "supplierTable", None)
            else:
                raise ValueError(f"Неизвестная таблица для экспорта: {table_name}")

            if table is None:
                raise AttributeError(f"Не удалось найти таблицу для экспорта в выбранной вкладке '{table_name}'.")

            # Собираем заголовки столбцов
            headers = []
            for col in range(table.columnCount()):
                header_item = table.horizontalHeaderItem(col)
                headers.append(header_item.text() if header_item else f"Column {col}")

            # Собираем данные из таблицы
            data = []
            for row in range(table.rowCount()):
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Создаем DataFrame и экспортируем в Excel
            df = pd.DataFrame(data, columns=headers)
            export_filename, _ = QFileDialog.getSaveFileName(self.main_window, "Экспорт данных", "", "Excel Files (*.xlsx)")
            if export_filename:
                df.to_excel(export_filename, index=False)

                # Настройка ширины столбцов с помощью openpyxl
                import openpyxl
                from openpyxl.utils import get_column_letter

                wb = openpyxl.load_workbook(export_filename)
                ws = wb.active

                for col in ws.columns:
                    max_length = 0
                    column = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass
                    adjusted_width = max_length + 2  # Добавляем небольшой отступ
                    ws.column_dimensions[column].width = adjusted_width

                wb.save(export_filename)

                QMessageBox.information(self.main_window, "Экспорт", f"Данные успешно экспортированы в {export_filename}")
                logging.info("Данные экспортированы в Excel успешно")
        except Exception as e:
            logging.error(f"Ошибка экспорта: {e}")
            QMessageBox.critical(self.main_window, "Ошибка экспорта", f"Ошибка экспорта: {str(e)}")

    def save_settings(self):
        """
        Сохраняет текущие настройки (интервал автообновления и режим изменения размера таблиц)
        в файл settings.json.
        """
        try:
            interval = self.refreshIntervalSpinBox.value()
            auto_resize = self.autoResizeCheckBox.isChecked()
            settings = {
                "refresh_interval": interval,
                "auto_resize": auto_resize
            }
            with open('settings.json', 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=4)
            QMessageBox.information(self.main_window, "Настройки", "Настройки успешно сохранены")
            logging.info("Настройки успешно сохранены")
        except Exception as e:
            logging.error(f"Ошибка сохранения настроек: {e}")
            QMessageBox.critical(self.main_window, "Ошибка", f"Ошибка сохранения настроек: {str(e)}")